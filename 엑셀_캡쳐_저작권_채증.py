import time

import pyautogui
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
from bs4 import BeautifulSoup
from 크롬드라이버_자동_업데이트 import chrome_driver

driver = chrome_driver(headless=False, image_enabled=True)
driver.set_window_position(0, 1)
driver.set_window_size(1000, 1000)


def load_excel():
    load_wb = load_workbook(r"C:\Users\user\OneDrive\바탕 화면\example.xlsx")  # 불러올 엑셀 경로
    load_ws = load_wb['Sheet1']  # 시트명
    data = load_ws.values
    columns = next(data)[0:]
    data_frame = pd.DataFrame(data, columns=columns)
    return load_wb, load_ws, data_frame


def screenshot(jpg):
    path = r"C:\Users\user\OneDrive\바탕 화면\image"  # 이미지를 저장할 경로
    pyautogui.screenshot(fr'{path}\{jpg}.png', region=(8, 8, 984, 984))
    now = datetime.now()
    print(f"{jpg} - 캡쳐 성공")
    return now


def naver_blog(url):
    driver.get(url)
    driver.implicitly_wait(10)
    response = driver.page_source
    soup = BeautifulSoup(response, 'html.parser')
    title = soup.select_one('meta[property="og:title"]')['content']
    write_date = soup.select_one('div.blog_authorArea > p').text
    return title, write_date


def naver_cafe(url):
    driver.get(url)
    driver.implicitly_wait(10)
    response = driver.page_source
    soup = BeautifulSoup(response, 'html.parser')
    title = soup.select_one('div.post_title > div.title_area > h2.tit').text
    write_date = soup.select_one('span.date.font_l').text.replace("작성일", "")
    return title, write_date


def main():
    load_wb, load_ws, data_frame = load_excel()

    # 첫 번째 행에서 필요한 정보들 골라오기
    for i, v in enumerate(zip(data_frame['URL'], data_frame['상단화면 파일명'], data_frame['게시물 제목'], data_frame['채증일'])):
        try:
            url, jpg, cnt_title, reg_date = v

            if cnt_title is not None:
                print("채증 됨")
                continue

            # 특정 구간에서 그만하고 싶거나 테스트하고 싶을 때 사용
            # if i < 5:
            #     break

            driver.get(url)
            driver.implicitly_wait(10)
            time.sleep(1)

            if reg_date is None:
                screenshot_time = screenshot(jpg)
                load_ws.cell(row=i + 2, column=6, value=screenshot_time)

            try:
                # 모바일 버전으로 들어가야 selector을 가져올 수 있음
                if url.find('blog.naver') != -1:
                    m_url = url.replace("blog.naver", "m.blog.naver")
                    title, write_date = naver_blog(m_url)

                elif url.find('cafe.naver') != -1:
                    m_url = url.replace("cafe.naver", "m.cafe.naver")
                    title, write_date = naver_cafe(m_url)

                else:
                    print("cafe나 blog 아님!!")
                    continue

            except Exception as e:
                print(e)
                continue

            load_ws.cell(row=i + 2, column=5, value=title)
            load_ws.cell(row=i + 2, column=7, value=write_date)

        except Exception as e:
            print(e)
            continue

    load_wb.save(r'C:\Users\user\OneDrive\바탕 화면\result.xlsx')
    driver.quit()


main()
